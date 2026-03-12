import pandas as pd
from openpyxl import load_workbook
import numpy as np


# Основная рабочая функция, которая открывает файл в режиме «только чтение» и построчно обходит все листы, возвращая данные порциями.
def read_lists_excel(file_path, chunk_size = 5000):
    # Открытие файла в режиме только чтения.
    wb = load_workbook(file_path, read_only = True)

    # Перебор листов
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Извлечения заголовка
        header_row = next(sheet.iter_rows(min_row = 2, values_only = True))
        header = [cell if cell is not None else f'Unnamed: {i}' for i, cell in enumerate(header_row)]

        chunk_index = 0
        rows_chunk = []

        # Чтение данных по блокам
        for row in sheet.iter_rows(max_row = 2, values_only = True):
            rows_chunk.append(row)
            if len(rows_chunk) == chunk_size:
                df = pd.DataFrame(rows_chunk, columns = header)
                yield sheet_name, chunk_index, df
                chunk_index += 1
                rows_chunk = []

        if rows_chunk:
            df = pd.DataFrame(rows_chunk, columns = header)
            yield sheet_name, chunk_index, df

    wb.close()

def data_type_recognition(df, sample_size=None, date_threshold=0.8, cat_max_unique=50, cat_ratio=0.05):
    if sample_size is not None and sample_size < len(df):
        df_sample = df.head(sample_size)
    else:
        df_sample = df

    result = {}

    for col in df_sample.columns:
        # Отбрасываем NaN – они не влияют на определение типа
        non_null = df_sample[col].dropna()
        if len(non_null) == 0:
            # Пустая колонка – считаем строковой (можно изменить на category_str, если нужно)
            result[col] = 'str'
            continue

        values = non_null.unique()
        total_non_null = len(non_null)

        # 1. Проверка на bool (True/False, 0/1, 'true'/'false')
        bool_set = {True, False, 1, 0, 'true', 'false', 'True', 'False', 'TRUE', 'FALSE'}
        if all(v in bool_set or (isinstance(v, str) and v.lower() in ('true', 'false')) for v in values):
            result[col] = 'bool'
            continue

        # 2. Проверка на числа (int / float)
        numeric_series = pd.to_numeric(pd.Series(values), errors='coerce')
        if not numeric_series.isna().any():
            # Все значения преобразовались в числа
            if np.all(numeric_series % 1 == 0):
                # Целые числа
                if len(values) <= cat_max_unique or (len(values) / total_non_null) <= cat_ratio:
                    result[col] = 'category_int'
                else:
                    result[col] = 'int'
                continue
            else:
                # Числа с плавающей точкой
                if len(values) <= cat_max_unique or (len(values) / total_non_null) <= cat_ratio:
                    result[col] = 'category_float'
                else:
                    result[col] = 'float'
                continue

        # 3. Проверка на datetime
        datetime_series = pd.to_datetime(pd.Series(values), errors='coerce')
        if not datetime_series.isna().all():
            success_rate = datetime_series.notna().mean()
            if success_rate >= date_threshold:
                if len(values) <= cat_max_unique or (len(values) / total_non_null) <= cat_ratio:
                    result[col] = 'category_datetime'
                else:
                    result[col] = 'datetime'
                continue

        # 4. Остальные случаи – строки или mixed
        types_in_col = set(type(v).__name__ for v in values)
        if len(types_in_col) > 1:
            # Разные типы Python (например, int и str) – mixed
            result[col] = 'mixed'
        else:
            # Однородные строки (или другие нераспознанные типы)
            if len(values) <= cat_max_unique or (len(values) / total_non_null) <= cat_ratio:
                result[col] = 'category_str'
            else:
                result[col] = 'str'

    return result