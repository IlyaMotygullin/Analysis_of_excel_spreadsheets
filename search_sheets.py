import os
import pandas as pd
from functools import lru_cache
from openpyxl import load_workbook

#Происходит поиск файла по всем папка ПК
def search_excel_file(filename, search_path='.'):
    for root, dirs, files in os.walk(search_path):
        if filename in files:
            full_path = os.path.join(root, filename)
            if check_excel_file(full_path):
                return full_path

    return None

#Проверка целостности файла, чисто для xlsx
def check_excel_file(filepath):
    try:
        wb = load_workbook(filename=filepath)
        #Проверка присутствия хотя бы одного листа
        if len(wb.sheetnames) > 0:
            return True
        else:
            return False
    except FileExistsError as e:
        print(e)
        return False

#Происходит кэширование файла
@lru_cache(maxsize=None)
def load_excel(file_path):
    #Приводим файл к абсолютному пути
    abs_path = os.path.abspath(os.path.normpath(file_path))
    return pd.ExcelFile(abs_path)

#Чтение файла
def read_excel(file_path, sheet_name=0, **kwargs):
    book = load_excel(file_path)
    return book.parse(sheet_name=sheet_name, **kwargs)